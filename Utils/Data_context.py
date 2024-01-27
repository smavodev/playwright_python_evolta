import csv
from openpyxl import load_workbook


def get_Data_from_CSV(file_path):
    data = []
    with open(file_path, 'r') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            data.append(row)
    return data


def get_Data_from_TXT(file_path):
    data = []
    with open(file_path, 'r') as txtfile:
        lines = txtfile.readlines()

        # Obtener los encabezados del archivo TXT (la primera línea)
        headers = lines[0].strip().split(',')

        # Iterar sobre las líneas y procesar los datos
        for line in lines[1:]:
            values = line.strip().split(',')
            # Crear un diccionario dinámicamente con claves y valores
            row_data = dict(zip(headers, values))
            data.append(row_data)

    return data


def get_Data_from_Excel(file_path):
    data = []
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Obtener los encabezados (nombres de columnas) desde la primera fila
    headers = [cell.value for cell in sheet[1]]

    # Iterar sobre las filas (empezando desde la segunda fila) y procesar los datos
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        data.append(row_data)

    return data